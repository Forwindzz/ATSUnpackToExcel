import json
import os
import yaml
import openpyxl
import utils
import json5
import re
from utils import *
from openpyxl.utils import get_column_letter
from collections import Counter
from openpyxl.styles import Font
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.formatting.rule import DataBarRule
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import time
from functools import partial

bold_font = Font(bold=True)
color_better_rule = ColorScaleRule(
    start_type='min', 
    start_color="FFFFFF",
    end_type='max', 
    end_color="00D131")

color_better_dark_rule = ColorScaleRule(
    start_type='min', 
    start_color="FFFFFF",
    end_type='max', 
    end_color="22b334")


color_worse_rule = ColorScaleRule(
    start_type='min', 
    start_color="00D131",
    end_type='max', 
    end_color="FFFFFF")

color_compare_rule = ColorScaleRule(
    start_type='min', 
    start_color="d80e4e",
    end_type='max', 
    end_color="61ef8b")

data_bar_rule = DataBarRule(
    start_type='min', end_type='max',
    color="5b86ec"
)

# 创建绿色填充样式
green_fill = PatternFill(start_color="b3ffca", end_color="b3ffca", fill_type="solid")

# 创建红色填充样式
red_fill = PatternFill(start_color="ff6b9a", end_color="ff6b9a", fill_type="solid")

rule_tick = CellIsRule(operator='equal', formula=['"√"'], fill=green_fill)
rule_cross = CellIsRule(operator='equal', formula=['"×"'], fill=red_fill)
group_hint_words = {} 
with open("translator_index.json5","r",encoding="utf-8") as f:
    group_hint_words = json5.load(f)

test_average_settings=[]
with open("test_average_settings.json5","r",encoding="utf-8") as f:
    test_average_settings = json5.load(f)

def get_trailing_number(s):
    m = re.search(r'\d+$', s)
    return int(m.group()) if m else None

def addTickCrossColor(sheet,rangeIndex):
    addEnumColor(sheet,rangeIndex,["b3ffca","ff6b9a"],{"√":"b3ffca","×":"ff6b9a"})

def addEnumColor(sheet,rangeIndex,colors=None,color_map=None):
    if colors is None:
        colors=["f7a9c2","f7bfa9","f7f0a9","d5f7a9","a9f7c3","a9f7ed","a9def7","a9baf7","b8a9f7","daa9f7","f7a9e8",
            "ef6d97","ef856d","eeb36f","eedf6f","c8ee6f","80ee6f","6feeb9","6fd5ee","6f90ee","9e6fee","e86fee",
            "6e2bdb","db2b86","db2b2b","db722b","dbbc2b","addb2b","69db2b","2bdb4e","2bdb9a","2bd9db","2b9fdb","2b73db","2b35db","d32bdb"
            ]
    if color_map is None:
        values=[]
        for cells in sheet[rangeIndex]:
            for cell in cells:
                values.append(cell.value)
        total=len(values)
        counter = Counter(values)
        enums = [item for item, count in counter.items() if count>=total/len(colors)*0.001]
        print("Enums count %d for %s>%s"%(len(enums),str(sheet),rangeIndex))
        color_map={}
        i=0
        for e in enums:
            if e is None or e=="":
                continue
            color = colors[i]
            color_map[e] = PatternFill(start_color=color, end_color=color, fill_type="solid")
            i = (i+1)%len(colors)
    else:
        target_mapping = {}
        for k,v in color_map.items():
            target_mapping[k]=PatternFill(start_color=v, end_color=v, fill_type="solid")
        color_map=target_mapping
    # excel has limitation of conditional format
    # we directly use format instead
    
    #for group, color in color_map.items():
    #    fill = PatternFill(start_color=color, end_color=color, fill_type=fill_type)
    #    rule = None
    #    if isinstance(group,str):
    #        rule = CellIsRule(operator='equal', formula=[f'"{group}"'], fill=fill)
    #    else:
    #        rule = CellIsRule(operator='equal', formula=[f'{group}'], fill=fill)
    #    sheet.conditional_formatting.add(rangeIndex, rule)
    for cells in sheet[rangeIndex]:
        for cell in cells:
            v = cell.value
            if v in color_map:
                cell.fill = color_map[v]

def bool2str(v):
    if v:
        return "√"
    else:
        return "×"

def getIndex(col:int,row:int):
    return "%s%d"%(get_column_letter(col),row)
def getIndexWithCol(col:str,row:int):
    return "%s%d"%(col,row)

removeSpiteRegx = re.compile(r'\（<spritename="[^"]*">\）')
removeColorRegx = re.compile(r'\<.*?\>')
numRegx = re.compile(r'(?<!\S)(?<!\{)[-+]?\d+%?(?!\})(?!\S)')
def translateLang(s):
    global globalSortedDisplayName
    if s is None:
        return ""
    s = str(s)
    nums = numRegx.findall(s)
    
    if len(nums)>0:
        s2=" "+s.replace(".asset"," ")+" ."
        nums = numRegx.findall(s2)
    for k,v in globalSortedDisplayName:
        s = s.replace(k,v)
    s = removeColorRegx.sub('', s)
    s = s.replace("_","").replace(" ","")
    if len(nums)>0:
        s+=" ("+",".join(nums)+")"
    return s.replace(".asset","")

def processBiomeModel(workbook:openpyxl.Workbook,biomeName, biomeInfo):
    displayName = translateLang(biomeName)
    # 概览部分：------------------------------------------------
    ov_sheet = workbook.create_sheet(("生态概览_%s"%(displayName))[0:31])
    ov_sheet.column_dimensions['A'].width = 20
    ov_sheet["A1"] = "地图类型概览 - %s"%displayName
    ov_sheet["C1"] = biomeName
    ov_sheet["A2"] = "> 基础部分"

    ov_sheet["A3"] = "地形视觉样式种类"
    ov_sheet["B3"] = len(biomeInfo["maps"])

    ov_sheet["A4"] = "允许使用邻近的负面正面修正" 
    ov_sheet["B4"] = bool2str(biomeInfo["canHaveModifier"]) 
    
    ov_sheet["A5"] = "基础奖励倍率" 
    ov_sheet["B5"] = biomeInfo["haseBaseReward"] 
    ov_sheet["A6"] = "基础周边修正奖励倍率" 
    ov_sheet["B6"] = biomeInfo["haseNearbyReward"] 
    
    ov_sheet["A8"] = "> 具体设定 " 
    ov_sheet["A9"] = "这部分会有很多配置名出现 " 

    ov_sheet["A10"] = "贸易路线时需求的订单种类" 
    ov_sheet["B10"] = len(biomeInfo["wantedGoods"]) 

    keys=["trade", "difficulty", "hostility","mapResources","gladesDeposits","gladesSprings","oreDeposits","gladesBuildings","gladesRelics","landPatches","newcomers","seasons"]
    displayKeys=["前来的商人配置","难度配置","敌意配置","树木配置","空地采集资源配置","空地泉水配置","空地矿石配置","空地装饰建筑配置","空地遗迹建筑配置","空地沃土配置","新村民来源配置","季节效果配置"]

    for i,key in enumerate(keys):
        row = 11+i
        ov_sheet["A%d"%row] = displayKeys[i]
        ov_sheet["B%d"%row] = biomeInfo[key]
    
    # 样式
    for cell in ov_sheet["A"]:
        cell.font = bold_font

def _genListName(strings:list):
    common_prefix = ""
    common_suffix = ""
    
    # 找到前缀和后缀
    for i in range(len(strings)):
        s=strings[i]
        if i==0:
            common_prefix = s
            common_suffix = s
        else:
            for i in range(len(common_prefix)-1,-1,-1):
                if common_prefix[i]!=s[i]:
                    common_prefix = common_prefix[:i]
                    break
            for j in range(len(common_suffix)-1,-1,-1):
                if common_suffix[-j]!=s[-j]:
                    common_suffix = common_suffix[-j:]
    breakI = None
    for i in range(len(common_prefix)-1,-1,-1):
        if common_prefix[i] in "0123456789":
            breakI=i
        else:
            break
    if breakI is not None:
        common_prefix = common_prefix[:breakI-1]
    breakI = None
    for i in range(len(common_suffix)-1,-1,-1):
        if common_suffix[-i] in "0123456789":
            breakI = i
        else:
            break
    if breakI is not None:
        common_suffix = common_suffix[-breakI+1:]
    dynamic_numbers = []
    for s in strings:
        match = re.search(f'{common_prefix}(\\d+){common_suffix}', s)
        if match:
            dynamic_numbers.append(int(match.group(1)))
    
    if len(dynamic_numbers)==0:
        result = f"{common_prefix}_{common_suffix}"
    else:
        result = f"{common_prefix}[{min(dynamic_numbers)}~{max(dynamic_numbers)}]{common_suffix}"
    return result

def _genGladeListName(glades:list):
    gs = []
    ws = []
    for glade in glades:
        ws.append(glade["weight"])
        gs.append(glade["glade"].replace(".json",""))
    name = _genListName(gs)#+"_%d_"%len(gs)
    gs.sort()
    ws.sort()
    name = name +"_"+hex(hash(tuple(gs+ws)))[3:7]
    return translateLang(name)

def processGenGladeModelIntegrate(workbook:openpyxl.Workbook,gladesGenGather,spGladeGenGather):
    gladesLists = {}
    gladesListsUser = {}
    rules=[]
    #将这些规则展平
    pc = ProgressCounter("GladeGen",len(gladesGenGather))
    for programName,gladeGenInfo in gladesGenGather.items():
        displayName = translateLang(programName)
        initGlades = [{"weight":100,"glade":g} for g in gladeGenInfo["initialGlades"]]
        name = _genGladeListName(initGlades)
        if name not in gladesLists:
            gladesLists[name]=initGlades
            gladesListsUser[name]=[]
        gladesListsUser[name].append(programName)
        rules.append(
            [
                programName,
                displayName,
                "初始空地",
                None,
                name,
                None,None,None,None, None,None,None,None, None,None,None,None
            ]
        )
        for i,levelInfo in enumerate(gladeGenInfo["levels"]):
            glades = levelInfo["glades"]
            name = _genGladeListName(glades)
            if name not in gladesLists:
                gladesLists[name]=glades
                gladesListsUser[name]=[]
            gladesListsUser[name].append(programName)
            rules.append(
                [
                    programName,
                    displayName,
                    "主空地",
                    i,
                    name,
                    levelInfo["parentOffset"]["x"],
                    levelInfo["parentOffset"]["y"],
                    levelInfo["randomXDeviation"]["x"],
                    levelInfo["randomXDeviation"]["y"],
                    levelInfo["randomYDeviation"]["x"],
                    levelInfo["randomYDeviation"]["y"],
                    levelInfo["amount"]["x"],
                    levelInfo["amount"]["y"],
                    levelInfo["angleRange"]["x"],
                    levelInfo["angleRange"]["y"],
                    levelInfo["useInitialAngles"],
                    "初始空地"
                ]
            )
        for i,extraIteration in enumerate(gladeGenInfo["extraIterations"]):
            levelInfo = extraIteration["iteration"]
            glades = levelInfo["glades"]
            name = _genGladeListName(glades)
            if name not in gladesLists:
                gladesLists[name]=glades
                gladesListsUser[name]=[]
            gladesListsUser[name].append(programName)
            rules.append(
                [
                    programName,
                    displayName,
                    "次空地",
                    i,
                    name,
                    levelInfo["parentOffset"]["x"],
                    levelInfo["parentOffset"]["y"],
                    levelInfo["randomXDeviation"]["x"],
                    levelInfo["randomXDeviation"]["y"],
                    levelInfo["randomYDeviation"]["x"],
                    levelInfo["randomYDeviation"]["y"],
                    levelInfo["amount"]["x"],
                    levelInfo["amount"]["y"],
                    levelInfo["angleRange"]["x"],
                    levelInfo["angleRange"]["y"],
                    levelInfo["useInitialAngles"],
                    "主空地%d"%extraIteration["parentLevel"]
                ]
            )
        pc.tick()
    pc.finish()
    pc = ProgressCounter("SP_GladeGen",len(spGladeGenGather))
    for programName, levelInfo in spGladeGenGather.items():
        displayName = translateLang(programName)
        glades = levelInfo["glades"]
        name = _genGladeListName(glades)
        if name not in gladesLists:
            gladesLists[name]=glades
            gladesListsUser[name]=[]
        gladesListsUser[name].append(programName)
        rules.append(
            [
                programName,
                displayName,
                "特殊空地",
                None,
                name,
                levelInfo["parentOffset"]["x"],
                levelInfo["parentOffset"]["y"],
                levelInfo["randomXDeviation"]["x"],
                levelInfo["randomXDeviation"]["y"],
                levelInfo["randomYDeviation"]["x"],
                levelInfo["randomYDeviation"]["y"],
                levelInfo["amount"]["x"],
                levelInfo["amount"]["y"],
                levelInfo["angleRange"]["x"],
                levelInfo["angleRange"]["y"],
                levelInfo["useInitialAngles"],
                "初始空地"
            ]
        )
        pc.tick()
    pc.finish()
    
    #表格：展示配置信息
    ov_sheet = workbook.create_sheet(("地图生成器-基础配置")[0:30])
    ov_sheet["A1"] = "地图生成配置"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 240
    ov_sheet["A2"] = """在你开始游戏时，程序会根据地图生态选取一个<生成配置>，然后根据<生成配置>，抽取一些空地，摆放他们：
    在<生成配置>中，会有3个<生成阶段>：
    预热阶段：<特殊空地> 会全部被生成，可以生成多个<特殊空地>的配置
    第一阶段：<初始空地> 选取1个初始空地
    第二阶段：<主要空地> 绕着<初始空地>生成一圈普通/危险/受禁空地，
    第三阶段：<次要空地> 将地图剩下的空隙用其他空地填补满，<次要空地>会绕着<主要空地>生成
    这个表格只包含地块摆放方式，如果想要它会选取那些地块，可以根据地块列表名去另一个表格查找。
    <配置程序/翻译名>：这个配置的名称，一般是生态群系，或者特殊修正的名字|<次序>：每个阶段会摆好几圈空地，这个就是顺序
    <空地抽签组>：一组预先准备好的空地，摆放的空地会从中选取，一共选取<最小空地数量>~<最大空地数量>个
    <使用初始角度>：我也不知道是啥|<围绕的空地>：围绕哪个空地进行摆放
    """
    
    nameList = [
        "生成配置程序名",
        "生成配置翻译名",
        "生成阶段",
        "次序",
        "空地抽签组",
        "最小距离（相对围绕的空地）",
        "最大距离（相对围绕的空地）",
        "x最小额外偏移",
        "x最大额外偏移",
        "y最小额外偏移",
        "y最大额外偏移",
        "最少空地数量",
        "最多空地数量",
        "最小空地间隔角度",
        "最大空地间隔角度",
        "使用初始角度？",
        "围绕的空地",
    ]
    width = [
        5,
        35,
        10,
        5,
        35,
        5,
        5,
        5,
        5,
        5,
        5,
        5,
        5,
        5,
        5,
        5,
        5,
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)

    dataLists = rules
    
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)
    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    frow = ov_sheet.max_row
    
    frow = ov_sheet.max_row
    ov_sheet.conditional_formatting.add('F4:F%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('G4:G%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('H4:H%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('I4:I%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('J4:J%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('K4:K%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('L4:L%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('M4:M%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('N4:N%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('O4:O%d'%frow, color_compare_rule)
    addEnumColor(ov_sheet,'B4:B%d'%frow)
    addEnumColor(ov_sheet,'C4:C%d'%frow)
    addEnumColor(ov_sheet,'E4:E%d'%frow)
    #ov_sheet.conditional_formatting.add('C4:C%d'%frow, color_better_rule)


    #表格：展示配置组
    ov_sheet = workbook.create_sheet(("地图生成器-抽签空地组")[0:30])
    ov_sheet["A1"] = "地图生成-用于抽签的一组组空地列表"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 20
    ov_sheet["A2"] = """
    """
    
    nameList = [
        "抽签空地组",
        "空地名",
        "权重",
        "组内概率"
    ]
    width = [
        25,
        35,
        10,
        10
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)

    dataLists = []
    for name,gladeLists in gladesLists.items():
        totalWeight = 0
        for gladeInfo in gladeLists:
            totalWeight+=gladeInfo["weight"]
        for gladeInfo in gladeLists:
            gladeInfo["chance"] = gladeInfo["weight"]/totalWeight
            dataLists.append(
                [
                    name,
                    translateLang(gladeInfo["glade"]),
                    gladeInfo["weight"],
                    gladeInfo["chance"]
                ]
            )
    
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)
    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    for cell in ov_sheet["D"]:
        cell.number_format = "0.00%"
    frow = ov_sheet.max_row
    ov_sheet.conditional_formatting.add('C4:C%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('D4:D%d'%frow, data_bar_rule)
    addEnumColor(ov_sheet,'A4:A%d'%frow)

    # 统计期望------------------------
    gladeGenToBiome = {}
    for biomeName,biomeInfo in biomesGather.items():
        if biomeName =="Capital.asset":
            continue
        difficultyName = biomeInfo["difficulty"]
        targetGenerationName=None
        for diff in difficultiesGather[difficultyName]:
            if diff["difficulty"] == "23 Ascension XX.asset":
                targetGenerationName = diff["generation"]
                break
        if targetGenerationName is None:
            print("Cannot find <23 Ascension XX> for %s > %s, use last difficulty instead"%(biomeName,difficultyName))
            targetGenerationName = difficultiesGather[difficultyName][-1]["generation"]
        if targetGenerationName not in gladeGenToBiome:
            gladeGenToBiome[targetGenerationName]=[]
        gladeGenToBiome[targetGenerationName].append(biomeName)
    
    for dic in test_average_settings:
        gen = dic["generation"]
        if gen not in gladeGenToBiome:
            gladeGenToBiome[gen]=[]
        gladeGenToBiome[gen]+=dic["biomeNames"]

    dataLists=[]
    for name,gladeLists in gladesLists.items():
        displayGladeListName = translateLang(name)
        users = gladesListsUser[name]
        biomeNames = set()
        for user in users:
            if user in gladeGenToBiome:
                for biomeName in gladeGenToBiome[user]:
                    biomeNames.add(biomeName)
        for biomeName in biomeNames:
            print("Simulate: %s - %s"%(biomeName,name))
            displayBiomeName = translateLang(biomeName)
            res={}
            def addRes(key,value,typeName,minV,maxV,chance,gladeType,uniqueImpact=False):
                if key not in res:
                    res[key]={
                        "value":0.0,
                        "typeName":typeName,
                        "min":minV,
                        "max":maxV,
                        "uniqueImpact":False,
                        "miss":1,
                        "missNormal":1,
                        "missDangerous":1,
                        "missForbidden":1,
                        "missOrigin":1,
                        "missSeal":1,
                        "glades":set([])
                        }
                rk = res[key]
                if rk["uniqueImpact"] or uniqueImpact:
                    rk["value"]=value
                    rk["miss"] = "1/该类型空地数量"
                    rk["uniqueImpact"] = True
                    return
                rk["value"]+=value
                rk["min"]=min(rk["min"],minV)
                rk["max"]=max(rk["max"],maxV)
                rk["uniqueImpact"] = rk["uniqueImpact"] or uniqueImpact
                rk["miss"] *= max(0,1-chance)
                rk["miss"+gladeType] *= max(0,1-chance)

            biomeInfo = biomesGather[biomeName]
            oreSettings = oresGenGather[biomeInfo["oreDeposits"]]
            springSettings = springsGatherDic[biomeInfo["gladesSprings"]]
            buildingSettings = buildingsGenGather[biomeInfo["gladesBuildings"]]
            relicSettings = relicGenGather[biomeInfo["gladesRelics"]]
            depositSettings = depositsGather[biomeInfo["gladesDeposits"]]
            for gladeTinyInfo in gladeLists:
                gladeChance = gladeTinyInfo["chance"]
                gladeName = gladeTinyInfo["glade"]
                gladeInfo = gladesGather[gladeName]
                gladeType = "Unknown"
                if "Origin" in gladeName:
                    gladeType = "Origin"
                elif "Dangerous" in gladeName:
                    gladeType = "Dangerous"
                elif "Forbidden" in gladeName:
                    gladeType = "Forbidden"
                elif "Normal" in gladeName or "Regular" in gladeName:
                    gladeType = "Normal"
                elif "Seal":
                    gladeType = "Seal"
                else:
                    print("Unknown Glade Name: "+gladeName)
                for relic in gladeInfo["relics"]:
                    lv = str(relic["Level"])
                    if lv not in relicSettings:
                        print("Unknown level %s for relics"%lv)
                        continue
                    for relicInfo in relicSettings[lv]:
                        chance = relicInfo["groupChance"] * gladeChance
                        if relicInfo["forceUniqueness"]!=0:
                            addRes(relicInfo["relic"],relicInfo["groupChance"],"遗迹",1,1,0,gladeType,True)
                        else:
                            if relicInfo["groupChance"]>=1:
                                addRes(relicInfo["relic"],chance,"遗迹",1,1,chance,gladeType)
                            else:
                                addRes(relicInfo["relic"],chance,"遗迹",0,1,chance,gladeType)
                for deposit in gladeInfo["deposits"]:
                    lv = str(deposit["Level"])
                    if lv not in depositSettings:
                        print("Unknown level %s for deposit"%lv)
                        continue
                    for depositInfo in depositSettings[lv]:
                        chance = gladeChance *0.5 *depositInfo["groupChance"]
                        average = (
                                deposit["AmountRange"]["x"]+
                                deposit["AmountRange"]["y"]+2
                            ) *0.5 * chance
                        addRes(depositInfo["deposit"],average,"采集资源",
                                deposit["AmountRange"]["x"]+1,
                                deposit["AmountRange"]["y"]+1,
                                chance,gladeType)
                for build in gladeInfo["buildings"]:
                    lv =str(build["Level"])
                    if lv not in buildingSettings:
                        print("Unknown level %s for buildings"%lv)
                        continue
                    for buildingInfo in buildingSettings[lv]:
                        chance = buildingInfo["groupChance"] * gladeChance
                        if buildingInfo["groupChance"]>=1:
                            addRes(buildingInfo["building"],chance,"装饰建筑",1,1,chance,gladeType)
                        else:
                            addRes(buildingInfo["building"],chance,"装饰建筑",0,1,chance,gladeType)
                for ore in gladeInfo["ore"]:
                    lv = ore["Level"]
                    flag=False
                    for oreLevel in oreSettings:
                        if lv==oreLevel["level"]:
                            for oreInfo in oreLevel["chances"]:
                                chance = oreInfo["groupChance"]*gladeChance
                                average = chance* (
                                        deposit["AmountRange"]["x"]+
                                        deposit["AmountRange"]["y"]+2
                                    )*0.5
                                addRes(oreInfo["ore"],average,"矿脉",
                                        deposit["AmountRange"]["x"]+1,
                                        deposit["AmountRange"]["y"]+1,
                                        chance,gladeType)
                            flag=True
                            break
                    if not flag:
                        print("Unknown level %d for ore"%lv)
                for spring in gladeInfo["springs"]:
                    lv = spring["Level"]
                    if lv not in springSettings:
                        print("Unknown level %d for spring"%lv)
                        continue
                    for springInfo in springSettings[lv]:
                        chance = springInfo["groupChance"] * gladeChance
                        if springInfo["groupChance"]>=1:
                            addRes(springInfo["spring"],chance,"泉水",1,1,chance,gladeType)
                        else:
                            addRes(springInfo["spring"],chance,"泉水",0,1,chance,gladeType)
                for grass in gladeInfo["landPatches"]:
                    v=grass["size"]["x"]*grass["size"]["y"]
                    addRes("沃土",gladeChance*v,"沃土",v,v,gladeChance,gladeType)
            for key,value in res.items():
                tag="未知"
                if value["missNormal"]<1:
                    tag = "小地"
                elif value["missDangerous"]<1:
                    tag = "危险地"
                elif value["missForbidden"]<1:
                    tag = "禁地"
                elif value["missOrigin"]<1:
                    tag = "初始地"
                elif value["missSeal"]<1:
                    tag = "封印地"
                dataLists.append(
                    [
                        displayGladeListName,
                        displayBiomeName,
                        value["typeName"],
                        translateLang(key),
                        value["value"],
                        value["miss"] if isinstance(value["miss"],str) else 1-value["miss"],
                        tag,
                        bool2str(value["uniqueImpact"])
                    ]
                )
    ov_sheet = workbook.create_sheet(("开空地的平均资源")[0:30])
    ov_sheet["A1"] = "这个表格可能最有用了"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 90
    ov_sheet["A2"] = """这边模拟了：如果你开一个空地，那么遭遇各种事件，获取各种资源的概率是多少？
    有的事件是全局唯一的（比如小空地宝鹿和商人事件，非诅咒林的部分闹鬼废墟。考古事件不在此列），
    他们会在开局时根据概率判定是否在这一局生成，如果生成，倾向于从最远的空地开始生成。
    小空地唯一事件在单局内100%加入生成，闹鬼建筑和幽灵则不是100%，具体看表格的平均数量，对于唯一建筑，这个表示为单局生成数量期望。
    另外遭遇概率指的是开这个类型的地遇到这个的概率，你可能遇到不止1个。
    沃土我没有计算被减去的部分，通常会减少1~2个沃土。
    """
    
    nameList = [
        "抽签空地组",
        "生物群系",
        "资源种类",
        "资源名称",
        "平均资源",
        "遭遇概率",
        "空地类型",
        "全局唯一生成"
    ]
    width = [
        25,
        20,
        12,
        25,
        12,
        12,
        18,
        10
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)
    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    for cell in ov_sheet["E"]:
        cell.number_format = "0.000"
    for cell in ov_sheet["F"]:
        cell.number_format = "0.00%"
    frow = ov_sheet.max_row
    i=4
    for i in range(4,frow+1):
        typeName = ov_sheet["C%d"%i].value
        vRange=(0,1)
        if typeName=="遗迹":
            vRange=(0,1)
        elif typeName=="采集资源":
            vRange=(0,3.0)
        elif typeName=="装饰建筑":
            vRange=(0,1.0)
        elif typeName=="矿脉":
            vRange=(0,2.7)
        elif typeName=="泉水":
            vRange=(0,1.0)
        elif typeName=="沃土":
            vRange=(0,25)
        
        intep = max(0,min(1,(ov_sheet["E%d"%i].value-vRange[0])/vRange[1]))
        color = interpolate_color("ffffff","b3ffca",intep)
        ov_sheet["E%d"%i].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
    
    ov_sheet.conditional_formatting.add('H4:H%d'%frow, rule_tick)
    ov_sheet.conditional_formatting.add('H4:H%d'%frow, rule_cross)
    ov_sheet.conditional_formatting.add('E4:E%d'%frow, data_bar_rule)
    ov_sheet.conditional_formatting.add('F4:F%d'%frow, data_bar_rule)
    #ov_sheet.conditional_formatting.add('D4:D%d'%frow, data_bar_rule)
    addEnumColor(ov_sheet,'A4:A%d'%frow)
    addEnumColor(ov_sheet,'B4:B%d'%frow)
    addEnumColor(ov_sheet,'C4:C%d'%frow)
    addEnumColor(ov_sheet,'D4:D%d'%frow)
    addEnumColor(ov_sheet,'G4:G%d'%frow)

def processGenGladeModel(workbook:openpyxl.Workbook, gladeGenName, gladeGen):
    displayName = translateLang(gladeGenName)


    # glade部分：---------------------------------------------
    ov_sheet = workbook.create_sheet(("空地生成规则_%s"%displayName)[0:30])
    ov_sheet.column_dimensions['A'].width = 25
    ov_sheet.column_dimensions['B'].width = 45
    ov_sheet.column_dimensions['D'].width = 45
    ov_sheet.column_dimensions['F'].width = 45
    ov_sheet.column_dimensions['H'].width = 45
    ov_sheet["A1"] = "地图生成-%s"%displayName
    ov_sheet["C1"] = gladeGenName
    ov_sheet["A2"] = "过程：先选一个<初始空地>生成，然后摆几圈<主要空地>，之后在<主要空地>周边添上<次要空地>，用于填补缝隙"
    ov_sheet["A3"] = "每次生成时，会在每一组选取若干个空地生成。具体的空地内容请查看另一个表格。生成时是顺(逆?)时针摆放空地，并且摆放时会有位置上的微小偏移"

    ov_sheet["A5"] = "初始空地列表"
    
    row = 6
    for i,v in enumerate(gladeGen["initialGlades"]):
        ov_sheet["B%d"%row]=translateLang(v)
        row+=1

    row+=1#gap
    ov_sheet["A%d"%row] = "<主要空地>列表"
    row+=1
    ov_sheet["A%d"%row] = ""
    ov_sheet["A%d"%(row+1)] = "到初始空地距离范围"
    ov_sheet["A%d"%(row+2)] = "x偏移范围"
    ov_sheet["A%d"%(row+3)] = "y偏移范围"
    ov_sheet["A%d"%(row+4)] = "空地间隔角度偏移范围"
    ov_sheet["A%d"%(row+5)] = "空地数量范围"
    
    startRow=row+7
    startCol = 2
    maxRow=0
    
    for i,level in enumerate(gladeGen["levels"]):
        curColName = get_column_letter(startCol+i*2)
        curCol2Name = get_column_letter(startCol+i*2+1)
        ov_sheet[getIndexWithCol(curColName,row+0)] = "主要空地组%d"%i
        ov_sheet[getIndexWithCol(curColName,row+1)] = level["parentOffset"]["x"]
        ov_sheet[getIndexWithCol(curCol2Name,row+1)] = level["parentOffset"]["y"]
        ov_sheet[getIndexWithCol(curColName,row+2)] = level["randomXDeviation"]["x"]
        ov_sheet[getIndexWithCol(curCol2Name,row+2)] = level["randomXDeviation"]["y"]
        ov_sheet[getIndexWithCol(curColName,row+3)] = level["randomYDeviation"]["x"]
        ov_sheet[getIndexWithCol(curCol2Name,row+3)] = level["randomYDeviation"]["y"]
        ov_sheet[getIndexWithCol(curColName,row+4)] = level["angleRange"]["x"]
        ov_sheet[getIndexWithCol(curCol2Name,row+4)] = level["angleRange"]["y"]
        ov_sheet[getIndexWithCol(curColName,row+5)] = level["amount"]["x"]
        ov_sheet[getIndexWithCol(curCol2Name,row+5)] = level["amount"]["y"]
        ov_sheet[getIndexWithCol(curColName,row+6)] = "空地名"
        ov_sheet[getIndexWithCol(curCol2Name,row+6)] = "权重"
        for j,gladeInfo in enumerate(level["glades"]):
            ov_sheet[getIndexWithCol(curColName,startRow+j)] = translateLang(gladeInfo["glade"])
            ov_sheet[getIndexWithCol(curCol2Name,startRow+j)] = gladeInfo["weight"]
        maxRow=max(maxRow,startRow+len(level["glades"]))
    for cell in ov_sheet[row+6]:
        cell.font = bold_font
    for cell in ov_sheet[row+0]:
        cell.font = bold_font

    row = maxRow+2
    ov_sheet["A%d"%row] = "<次要空地>列表"
    row+=1
    ov_sheet["A%d"%row] = ""
    ov_sheet["A%d"%(row+1)] = "到初始空地距离范围"
    ov_sheet["A%d"%(row+2)] = "x偏移范围"
    ov_sheet["A%d"%(row+3)] = "y偏移范围"
    ov_sheet["A%d"%(row+4)] = "空地间隔角度偏移范围"
    ov_sheet["A%d"%(row+5)] = "空地数量范围"
    ov_sheet["A%d"%(row+6)] = "围绕哪个组别的主空地"

    startRow=row+8
    startCol = 2
    
    for i,levelBase in enumerate(gladeGen["extraIterations"]):
        level = levelBase["iteration"]
        curColName = get_column_letter(startCol+i*2)
        curCol2Name = get_column_letter(startCol+i*2+1)
        ov_sheet[getIndexWithCol(curColName,row+0)] = "次要空地组%d"%i
        ov_sheet[getIndexWithCol(curColName,row+1)] = level["parentOffset"]["x"]
        ov_sheet[getIndexWithCol(curCol2Name,row+1)] = level["parentOffset"]["y"]
        ov_sheet[getIndexWithCol(curColName,row+2)] = level["randomXDeviation"]["x"]
        ov_sheet[getIndexWithCol(curCol2Name,row+2)] = level["randomXDeviation"]["y"]
        ov_sheet[getIndexWithCol(curColName,row+3)] = level["randomYDeviation"]["x"]
        ov_sheet[getIndexWithCol(curCol2Name,row+3)] = level["randomYDeviation"]["y"]
        ov_sheet[getIndexWithCol(curColName,row+4)] = level["angleRange"]["x"]
        ov_sheet[getIndexWithCol(curCol2Name,row+4)] = level["angleRange"]["y"]
        ov_sheet[getIndexWithCol(curColName,row+5)] = level["amount"]["x"]
        ov_sheet[getIndexWithCol(curCol2Name,row+5)] = level["amount"]["y"]
        ov_sheet[getIndexWithCol(curColName,row+6)] = "主要空地组%d"%levelBase["parentLevel"]
        ov_sheet[getIndexWithCol(curColName,row+7)] = "空地名"
        ov_sheet[getIndexWithCol(curCol2Name,row+7)] = "权重"
        for j,gladeInfo in enumerate(level["glades"]):
            ov_sheet[getIndexWithCol(curColName,startRow+j)] = translateLang(gladeInfo["glade"])
            ov_sheet[getIndexWithCol(curCol2Name,startRow+j)] = gladeInfo["weight"]
        maxRow=max(maxRow,startRow+len(level["glades"]))
    for cell in ov_sheet[row+7]:
        cell.font = bold_font
    for cell in ov_sheet[row+0]:
        cell.font = bold_font
    # 样式
    for cell in ov_sheet["A"]:
        cell.font = bold_font
    pass

def _gladeType(glade):
    info = ""
    if glade["as_init_glade"]>0:
        info+="初-"
    if glade["as_main_glade"]>0:
        info+="主-"
    if glade["as_sub_glade"]>0:
        info+="次-"
    if glade["as_sp_glade"]>0:
        info+="特殊-"
    return info[:-1]

def _gladeGrassCount(glade):
    count=0
    for landPatch in glade["landPatches"]:
        count += landPatch["size"]["x"]*landPatch["size"]["y"]
    return count

def _minMaxGladeDeposite(glade):
    countMin=0
    countMax=0
    for deposit in glade["deposits"]:
        countMin += deposit["AmountRange"]["x"]+1
        countMax += deposit["AmountRange"]["y"]+1
    return countMin,countMax

def _minMaxOreDeposite(glade):
    countMin=0
    countMax=0
    for ore in glade["ore"]:
        countMin += ore["AmountRange"]["x"]+1
        countMax += ore["AmountRange"]["y"]+1
    return countMin,countMax

def __genGladeGroupKey(k):
    if "遗迹" in k:
        return 0
    if "资源" in k:
        return 1
    if "建筑" in k:
        return 2
    if "矿石" in k:
        return 3
    if "泉水" in k:
        return 4
    return 5
def _genGladeGroup(glade):
    result = []
    for relic in glade["relics"]:
        result.append("遗迹组%d"%relic["Level"])
    for deposit in glade["deposits"]:
        result.append("资源组%d"%deposit["Level"])
    for build in glade["buildings"]:
        result.append("建筑组%d"%build["Level"])
    for ore in glade["ore"]:
        result.append("矿石组%d"%ore["Level"])
    for spring in glade["springs"]:
        result.append("泉水组%d"%spring["Level"])
    counter = Counter(result)
    result_list = [f"{count}×{item}" if count > 1 else item for item, count in counter.items()]
    result_list.sort(key=__genGladeGroupKey)
    return result_list

def _genGladeDetailInfo(glade):
    result=[]
    for relic in glade["relics"]:
        result.append(
            "遗迹组%d(%d,%d>%d°)"%(
                relic["Level"],
                relic["Field"]["x"],relic["Field"]["y"],
                relic["Rotation"] if "Rotation" in relic else 0
                )
            )
    for deposit in glade["deposits"]:
        result.append(
            "资源组%d(%d,%d,[%d,%d])"%(
                deposit["Level"],
                deposit["Field"]["x"],deposit["Field"]["y"],
                deposit["AmountRange"]["x"],deposit["AmountRange"]["y"]
                )
            )
    for build in glade["buildings"]:
        result.append(
            "建筑组%d(%d,%d>%d°)"%(
                build["Level"],
                build["Field"]["x"],build["Field"]["y"],
                build["Rotation"] if "Rotation" in build else 0
                )
            )
    for ore in glade["ore"]:
        result.append(
            "资源组%d(%d,%d,[%d,%d])"%(
                ore["Level"],
                ore["Field"]["x"],ore["Field"]["y"],
                ore["AmountRange"]["x"],ore["AmountRange"]["y"]
                )
            )
    for spring in glade["springs"]:
        result.append(
            "泉水组%d(%d,%d)"%(
                spring["Level"],
                spring["Field"]["x"],spring["Field"]["y"]
                )
            )
    for grass in glade["landPatches"]:
        result.append(
            "沃土组(%d,%d,[%d,%d])"%(
                grass["Field"]["x"],grass["Field"]["y"],
                grass["size"]["x"],grass["size"]["y"]
                )
            )
    if "isOrigin" in glade and glade["isOrigin"]:
        result.append("是初始")
    if "Field" in glade["origin"]:
        result.append("初始点(%d,%d)"%(glade["origin"]["Field"]["x"],glade["origin"]["Field"]["y"]))
    if "hearth" in glade and glade["hearth"] is not None:
        result.append("有主火塘")
    if "storage" in glade and glade["storage"] is not None:
        result.append("有主仓库")
    return result

def _getGroupHintWords(types,index):
    index = str(index)
    if index in group_hint_words[types]:
        return group_hint_words[types][index]
    else:
        return types+index
def _genGladeTranslateIndexGroup(group_list):
    results=[]
    for v in group_list:
        dic = None
        if "遗迹" in v:
            dic = group_hint_words["relics"]
        elif "资源" in v:
            dic = group_hint_words["deposits"]
        elif "建筑" in v:
            dic = group_hint_words["buildings"]
        elif "矿石" in v:
            dic = group_hint_words["ores"]
        elif "泉水" in v:
            dic = group_hint_words["springs"]
        else:
            continue
        index=str(get_trailing_number(v))
        new_v=None
        if index in dic:
            new_v = dic[index]
        else:
            print("Not comment: "+v)
            new_v = v
        if "×" in v:
            v = v[:v.find("×")]
            new_v = v+"×"+new_v
        results.append(new_v)
    return results

def _addDataToWorkSheet(dataLists, worksheet,startCol=1,startRow=1):
    for i,dataLine in enumerate(dataLists):
        for j,data in enumerate(dataLine):
            worksheet[getIndex(startCol+j,startRow+i)] = data

def _addTitleBar(nameList,width,worksheet,startCol=1,startRow=1):
    for i in range(len(nameList)):
        worksheet[getIndex(startCol+i,startRow)] = nameList[i]
    for i in range(len(width)):
        worksheet.column_dimensions[get_column_letter(i+1)].width = width[i]

def processGlade(workbook:openpyxl.Workbook, glades):
    ov_sheet = workbook.create_sheet(("空地模板")[0:30])
    ov_sheet["A1"] = "空地模板"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 100
    ov_sheet["A2"] = """所有空地生成都是基于这些模板，表格只展示重要部分。
    这些模板决定了这个空地有哪些类型的资源，摆放在哪里，数量为多少，以及摆放哪些生成组。
    生成组会决定资源的类型，会从一组候选选择中抽取一个决定，生成组在不同生物群系的地图下会有所不同。
    空地类别：影响生成方式，一般靠近古老火塘的为主空地，主空地外围为次要空地。
    累计xx资源：在生成1个资源点后，周边追加[min,max]个资源点，我们统计累计数值
    生成候选组：这里会告知这个空地能会摆放哪些组别的资源节点，箱子，事件建筑，废墟等等...
    候选组具体内容：xx组编号(坐标x,坐标y>旋转标识°,[最小额外值或沃土长，最大额外值或沃土宽])
    """
    
    nameList = [
        "空地程序名",
        "空地翻译名",
        "长度",
        "宽度",
        "类别",
        "总权重",
        "空地泉水资源数量",
        "沃土候选生成数量",
        "资源点集群数",
        "采集资源最大累计数量",
        "采集资源最小累计数量",
        "矿石最大累计数量",
        "矿石最小累计数量",
        "装饰建筑数量",
        "遗迹建筑数量",
        "生成候选组（候选组的内容查看其他表格）",
        "候选组具体内容"
    ]
    width = [
        5,
        25,
        5,
        5,
        7,
        6,
        5,
        5,
        5,
        5,
        5,
        5,
        5,
        5,
        5,
        60,
        120
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)

    dataLists = []
    pc = ProgressCounter("Glade",len(gladesGenGather))
    for gladeName, glade in glades.items():
        minDeposit, maxDeposit = _minMaxGladeDeposite(glade)
        minOre, maxOre = _minMaxOreDeposite(glade)
        groups = _genGladeGroup(glade)
        groupTranslate = _genGladeTranslateIndexGroup(groups)
        data = [
                gladeName,
                translateLang(gladeName),
                glade["size"]["y"],
                glade["size"]["x"],
                _gladeType(glade),
                glade["total_weight"],
                len(glade["springs"]),
                _gladeGrassCount(glade),
                len(glade["deposits"]),
                minDeposit,
                maxDeposit,
                minOre,
                maxOre,
                len(glade["buildings"]),
                len(glade["relics"]),
                "，".join(groupTranslate),
                "，".join(_genGladeDetailInfo(glade))
                
            ]
        dataLists.append(data)
        pc.tick()
    pc.finish()
    
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)
    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    frow = ov_sheet.max_row
    ov_sheet.conditional_formatting.add('C4:C%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('D4:D%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('F4:F%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('G4:G%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('H4:H%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('I4:I%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('J4:J%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('K4:K%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('L4:L%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('M4:M%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('N4:N%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('O4:O%d'%frow, color_better_rule)
    
    #addEnumColor(ov_sheet,'E4:E%d'%frow)
    #addEnumColor(ov_sheet,'P4:P%d'%frow)
    # 其他表格：遗迹组 
    ov_sheet = workbook.create_sheet(("遗迹组")[0:30])
    ov_sheet["A1"] = "遗迹建筑模板 - 箱子/危险受禁事件建筑/废墟建筑"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 120
    ov_sheet["A2"] = """2×遗迹组10 表示这个glade会生成2个遗迹，遗迹的类型通过抽签决定，会在相应<生态环境配置>的<组别10>中加权抽签
抽取的概率由权重决定。
被标识为全局唯一的建筑，一场游戏内最多生成1次。
全局生成概率为0~1，如果判定没有成功生成，那么整局游戏内你都不会看到这个建筑。如果判定成功，那么它必定最先出现在最后放置的空地上。
        """
    
    nameList = [
        "配置程序名",
        "配置翻译名",
        "组别编号",
        "组别名称",
        "全局唯一",
        "全局生成概率",
        "权重",
        "组内概率",
        "建筑名称翻译",
        "建筑名称",
    ]
    width = [
        5,
        25,
        7,
        18,
        7,
        7,
        10,
        10,
        30,
        25
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    dataLists = []
    pc = ProgressCounter("Relics",len(relicGenGather))
    for typeName,indexDict in relicGenGather.items():
        indexList = [(int(index),relicList) for index,relicList in indexDict.items()]
        indexList = sorted(indexList, key=lambda x:x[0])
        for index, relicList in indexList:
            #totalWeight = 0
            #for relic in relicList:
            #    totalWeight+=relic["weight"]
            for relic in relicList:
                data=[
                    typeName,
                    translateLang(typeName),
                    index,
                    _getGroupHintWords("relics",index),
                    bool2str(relic["forceUniqueness"]),
                    relic["chance"],
                    relic["weight"],
                    relic["groupChance"],#relic["weight"]/totalWeight,
                    translateLang(relic["relic"]),
                    relic["relic"]
                ]
                dataLists.append(data)
        pc.tick()
    pc.finish()
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)
    
    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    for cell in ov_sheet["H"]:
        cell.number_format = "0.00%"
    frow = ov_sheet.max_row
    ov_sheet.conditional_formatting.add('G4:G%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('F4:F%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('H4:H%d'%frow, data_bar_rule)
    ov_sheet.conditional_formatting.add('E4:E%d'%frow, rule_tick)
    ov_sheet.conditional_formatting.add('E4:E%d'%frow, rule_cross)
    addEnumColor(ov_sheet,'C4:C%d'%frow)
    addEnumColor(ov_sheet,'D4:D%d'%frow)
    addEnumColor(ov_sheet,'B4:B%d'%frow)


    # 其他表格：资源采集组 
    ov_sheet = workbook.create_sheet(("资源采集组")[0:30])
    ov_sheet["A1"] = "可采集资源点模板"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 70
    ov_sheet["A2"] = """2×资源组10 表示这个空地会生成2个资源节点，资源节点的类型通过抽签决定，会在相应<生态环境配置>的<组别10>中加权抽签
抽取的概率由权重决定。
决定资源节点类型后，还会在这个资源节点旁边生成若干个相同的资源节点，比如小资源点通常会额外生成3个左右。
这个额外资源节点数量由空地模板决定（也就是空地模板里的累计最大最小资源节点数量）
        """
    
    nameList = [
        "配置程序名",
        "配置翻译名",
        "组别编号",
        "组别名称",
        "权重",
        "组内概率",
        "节点名称翻译",
        "节点名称",
    ]
    width = [
        5,
        25,
        7,
        18,
        7,
        10,
        30,
        25
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    dataLists = []
    for typeName,indexDict in depositsGather.items():
        indexList = [(int(index),depositList) for index,depositList in indexDict.items()]
        indexList = sorted(indexList, key=lambda x:x[0])
        for index, depositList in indexList:
            #totalWeight = 0
            #for deposit in depositList:
            #    totalWeight+=deposit["amount"]
            for deposit in depositList:
                data=[
                    typeName,
                    translateLang(typeName),
                    index,
                    _getGroupHintWords("deposits",index),
                    deposit["amount"],
                    deposit["groupChance"],#deposit["amount"]/totalWeight,
                    translateLang(deposit["deposit"]),
                    deposit["deposit"]
                ]
                dataLists.append(data)
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)
    
    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    for cell in ov_sheet["F"]:
        cell.number_format = "0.00%"
    frow = ov_sheet.max_row
    ov_sheet.conditional_formatting.add('E4:E%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('F4:F%d'%frow, data_bar_rule)
    addEnumColor(ov_sheet,'C4:C%d'%frow)
    addEnumColor(ov_sheet,'D4:D%d'%frow)
    addEnumColor(ov_sheet,'B4:B%d'%frow)
    
    # 其他表格：遗迹建筑组
    ov_sheet = workbook.create_sheet(("建筑组")[0:30])
    ov_sheet["A1"] = "建筑组模板"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 70
    ov_sheet["A2"] = """2×建筑组10 表示这个空地会生成2个建筑，建筑的类型通过抽签决定，会在相应<生态环境配置>的<组别10>中加权抽签
抽取的概率由权重决定。
需要村民处理事件的建筑属于<遗迹组>。这里的建筑组开地即用。
        """
    
    nameList = [
        "配置程序名",
        "配置翻译名",
        "组别编号",
        "组别名称",
        "权重",
        "组内概率",
        "节点名称翻译",
        "节点名称",
    ]
    width = [
        5,
        25,
        7,
        18,
        7,
        10,
        30,
        25
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    dataLists = []
    for typeName,indexDict in buildingsGenGather.items():
        indexList = [(int(index),buildingList) for index,buildingList in indexDict.items()]
        indexList = sorted(indexList, key=lambda x:x[0])
        for index, buildingList in indexList:
            totalWeight = 0
            for building in buildingList:
                totalWeight+=building["weight"]
            for building in buildingList:
                data=[
                    typeName,
                    translateLang(typeName),
                    index,
                    _getGroupHintWords("buildings",index),
                    building["weight"],
                    building["weight"]/totalWeight,
                    translateLang(building["building"]),
                    building["building"]
                ]
                dataLists.append(data)
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)

    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    for cell in ov_sheet["F"]:
        cell.number_format = "0.00%"
    frow = ov_sheet.max_row
    ov_sheet.conditional_formatting.add('E4:E%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('F4:F%d'%frow, data_bar_rule)
    addEnumColor(ov_sheet,'C4:C%d'%frow)
    addEnumColor(ov_sheet,'D4:D%d'%frow)
    addEnumColor(ov_sheet,'B4:B%d'%frow)
    

    # 其他表格：矿石组
    ov_sheet = workbook.create_sheet(("矿石组")[0:30])
    ov_sheet["A1"] = "矿石组模板"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 70
    ov_sheet["A2"] = """2×矿石组1 表示这个空地会生成2个矿脉，矿脉的类型通过抽签决定，会在相应<生态环境配置>的<组别10>中加权抽签
抽取的概率由权重决定。
生成时，会先生成中心位置的矿物资源点，然后周围按照额外矿脉的配置，增加矿脉资源点
        """
    
    nameList = [
        "配置程序名",
        "配置翻译名",
        "组别编号",
        "组别名称",
        "权重",
        "组内概率",
        "节点名称翻译",
        "节点名称",
    ]
    width = [
        5,
        25,
        7,
        18,
        7,
        10,
        30,
        25
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    dataLists = []
    pc = ProgressCounter("Ore",len(oresGenGather))
    for typeName,indexListInfo in oresGenGather.items():
        indexList = [(int(content["level"]),content["chances"]) for content in indexListInfo]
        indexList = sorted(indexList, key=lambda x:x[0])
        for index, oreList in indexList:
            totalWeight = 0
            for ore in oreList:
                totalWeight+=ore["amount"]
            for ore in oreList:
                data=[
                    typeName,
                    translateLang(typeName),
                    index,
                    _getGroupHintWords("ores",index),
                    ore["amount"],
                    ore["amount"]/totalWeight,
                    translateLang(ore["ore"]),
                    ore["ore"]
                ]
                dataLists.append(data)
        pc.tick()
    pc.finish()
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)

    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    for cell in ov_sheet["F"]:
        cell.number_format = "0.00%"
    frow = ov_sheet.max_row
    ov_sheet.conditional_formatting.add('E4:E%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('F4:F%d'%frow, data_bar_rule)
    addEnumColor(ov_sheet,'C4:C%d'%frow)
    addEnumColor(ov_sheet,'D4:D%d'%frow)
    addEnumColor(ov_sheet,'B4:B%d'%frow)

    # 其他表格：泉水组
    ov_sheet = workbook.create_sheet(("泉水组")[0:30])
    ov_sheet["A1"] = "泉水组模板"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 70
    ov_sheet["A2"] = """泉水组1 表示这个空地会生成1个泉水，泉水的类型通过抽签决定，会在相应<生态环境配置>的<组别10>中加权抽签
抽取的概率由权重决定。
        """
    
    nameList = [
        "配置程序名",
        "配置翻译名",
        "组别编号",
        "组别名称",
        "权重",
        "组内概率",
        "节点名称翻译",
        "节点名称",
    ]
    width = [
        5,
        25,
        7,
        18,
        7,
        10,
        30,
        25
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    dataLists = []
    pc = ProgressCounter("Building",len(springsGenGather))
    for typeName,listInfo in springsGenGather.items():
        dic={}
        for info in listInfo:
            lv = info["level"]
            if lv not in dic:
                dic[lv]=[]
            dic[lv].append(info)
        indexList = [(lv,listValues) for lv,listValues in dic.items()]
        indexList = sorted(indexList, key=lambda x:x[0])
        for index, springList in indexList:
            #totalWeight = 0
            #for spring in springList:
            #    totalWeight+=spring["weight"]
            for spring in springList:
                data=[
                    typeName,
                    translateLang(typeName),
                    index,
                    _getGroupHintWords("springs",index),
                    spring["weight"],
                    spring["groupChance"],#spring["weight"]/totalWeight,
                    translateLang(spring["spring"]),
                    spring["spring"]
                ]
                dataLists.append(data)
        pc.tick()
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)
    pc.finish()
    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    for cell in ov_sheet["F"]:
        cell.number_format = "0.00%"
    frow = ov_sheet.max_row
    ov_sheet.conditional_formatting.add('E4:E%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('F4:F%d'%frow, data_bar_rule)
    addEnumColor(ov_sheet,'C4:C%d'%frow)
    addEnumColor(ov_sheet,'D4:D%d'%frow)
    addEnumColor(ov_sheet,'B4:B%d'%frow)




def GenEventSheet(workbook:openpyxl.Workbook):

    # relic events
    dataLists = []
    for relicAssetName, relicInfo in relicsGather.items():

        dataLists.append(
            [
                relicAssetName,
                translateLang(relicAssetName),
            ]
        )
        pass

    ov_sheet = workbook.create_sheet(("处理空地事件")[0:30])
    ov_sheet["A1"] = "所有森林空地事件"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 70
    ov_sheet["A2"] = """
        """
    
    nameList = [
        "程序名称",
        "翻译名称",
        "描述",
        "处理人数",
        "难度环境",
        "抉择名称",
        "抉择类型",
        "抉择奖励",
        "工作时间",
        "需求物品组",
        "处理时承受的效果"
    ]
    width = [
        5,
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)

    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    #for cell in ov_sheet["F"]:
    #    cell.number_format = "0.00%"
    #frow = ov_sheet.max_row
    #ov_sheet.conditional_formatting.add('E4:E%d'%frow, color_compare_rule)
    #addEnumColor(ov_sheet,'C4:C%d'%frow)
        
    # ............................................
    ov_sheet = workbook.create_sheet(("过期空地事件")[0:30])
    ov_sheet["A1"] = "所有森林空地事件"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 70
    ov_sheet["A2"] = """
        """
    
    nameList = [
        "程序名称",
        "翻译名称",
        "宽x",
        "长y"
        "过期时间（从发现开始计算）",
        "永久效果",
        "出现时效果",
        "过期效果"
    ]
    width = [
        5,
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)

    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    #for cell in ov_sheet["F"]:
    #    cell.number_format = "0.00%"
    #frow = ov_sheet.max_row
    #ov_sheet.conditional_formatting.add('E4:E%d'%frow, color_compare_rule)
    #addEnumColor(ov_sheet,'C4:C%d'%frow)

def tryAdd(asset, key,postProcess= lambda x:x,defaultValue=None):
        if key in asset:
            return postProcess(asset[key])
        try:
            return postProcess(defaultValue)
        except:
            return defaultValue

def filterNoneToStr(obj):
    return obj if obj is not None else ""

def GenGoodSheet(workbook:openpyxl.Workbook):
    # goods
    dataLists = []
    amberPrice = goodsGather["_Valuable_ Amber.asset"]["tradingSellValue"]
    pc = ProgressCounter("Goods",len(buildingsGather))
    for goodAssetName, goodAsset in goodsGather.items():
        dataLists.append([
            goodAsset["m_Name"],
            translateLang(goodAsset["displayName"]["key"]),
            goodAsset["consoleId"],
            translateLang(goodAsset["category"]),
            bool2str(goodAsset["eatable"]),
            goodAsset["eatingFullness"],
            bool2str(goodAsset["canBeBurned"]),
            goodAsset["burningTime"],
            goodAsset["tradingSellValue"],
            goodAsset["tradingBuyValue"],
            goodAsset["tradingSellValue"]/goodAsset["tradingBuyValue"],
            goodAsset["tradingSellValue"]/amberPrice,
            ",".join([translateLang(x) for x in goodAsset["tags"]]),
            goodAsset["order"],
            translateLang(goodAsset["shortDescription"]["key"]),
            translateLang(goodAsset["description"]["key"])

        ])
        pc.tick()
        pass
    pc.finish()

    ov_sheet = workbook.create_sheet(("物品")[0:30])
    ov_sheet["A1"] = "物品"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 90
    ov_sheet["A2"] = """总之就是非常常见的物品，注意[火花露]在当前版本被移除
        <命令名称>：启用控制台的时候可以用（测试版）
        <恢复饱食>：会恢复相应的饥饿耐受度
        <价格>：这里的价格是绝对价值，在游戏中需要除以琥珀的出售价值来计算价格
        <价格比例>：出售价格打了多少折，出售价格/购买价格
        <顺序标号>：影响显示顺序...（好像没什么用）
        """
    
    nameList = [
        "程序名称（Unity Prefab名称）",
        "物品名称",
        "命令名称",
        "类别",
        "能吃",
        "恢复饱食",
        "能烧",
        "燃烧时间",
        "出售价格",
        "购买价格",
        "价格比例",
        "琥珀价格",
        "额外标签",
        "顺序标号",
        "简短描述",
        "描述"
    ]
    width = [
        5,
        15,
        10,
        18,
        5,
        8,
        5,
        8,
        8,
        8,
        8,
        8,
        15,
        5,
        25,
        50
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)

    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    for cell in ov_sheet["I"]:
        cell.number_format = "0.00"
    for cell in ov_sheet["J"]:
        cell.number_format = "0.00"
    for cell in ov_sheet["K"]:
        cell.number_format = "0%"
    for cell in ov_sheet["L"]:
        cell.number_format = "0.00"
    frow = ov_sheet.max_row
    ov_sheet.conditional_formatting.add('E4:E%d'%frow, rule_tick)
    ov_sheet.conditional_formatting.add('E4:E%d'%frow, rule_cross)
    ov_sheet.conditional_formatting.add('G4:G%d'%frow, rule_tick)
    ov_sheet.conditional_formatting.add('G4:G%d'%frow, rule_cross)
    ov_sheet.conditional_formatting.add('F4:F%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('H4:H%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('I4:I%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('J4:J%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('K4:K%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('L4:L%d'%frow, data_bar_rule)
    addEnumColor(ov_sheet,'D4:D%d'%frow)
    print("Generated good, row to %d"%frow)
    #--------------------------------------------

    def mapGrade(s:str):
        if s=="Grade0.asset":
            return 0
        elif s=="Grade1.asset":
            return 1
        elif s=="Grade3.asset":
            return 3
        elif s=="Grade3.asset":
            return 3
        else:
            return int(s.replace("Grade","").replace(".asset",""))
        

    def mapGradeStr(lv:int):
        return ["☆","★","★★","★★★","★★★★","★★★★★"][lv]
    
    def goodInfoStr(goodInfo):
        a = goodInfo["amount"]
        if a==0:
            return ""
        return str(a) + " × " + translateLang(goodInfo["good"])
    
    def goodsInfoStr(goodsInfo, d="\n"):
        s = ""
        for good in goodsInfo:
            s+=goodInfoStr(good)+d
        return s[:-1]
    
    dataLists=[]
    recipeNames={}
    linesHeightInfo = []
    
    
    
    def recipeTypeStr(s):
        maps={
            "InstitutionRecipeModel.cs":"服务",
            "GathererHutRecipeModel.cs":"采集",
            "FishingHutRecipeModel.cs":"鱼塘",
            "CampRecipeModel.cs":"伐木",
            "CollectorRecipeModel.cs":"收集",
            "WorkshopRecipeModel.cs":"生产",
            "FarmRecipeModel.cs":"种植",
            "MineRecipeModel.cs":"挖矿",
            "RainCatcherRecipeModel.cs":"集雨"
        }
        return maps[s]
    
    pc = ProgressCounter("Recipe",len(recipesGather))
    for recipeAssetName, recipeAsset in recipesGather.items():
        #if recipeAsset["m_Script"] not in ["CollectorRecipeModel.cs","WorkshopRecipeModel.cs","RainCatcherRecipeModel.cs"]:
        #    recipeNames[recipeAssetName] = translateLang(recipeAssetName)
        #    continue
        lines=1
        grade = mapGrade(translateLang(recipeAsset["grade"]))
        data = [
            recipeAsset["m_Name"],
            translateLang(recipeAsset["m_Name"]),
            recipeTypeStr(recipeAsset["m_Script"]),
            mapGradeStr(grade)
        ]
        data.append(tryAdd(recipeAsset,"productionTime"))
        if "producedGood" in recipeAsset:
            data.append(translateLang(recipeAsset["producedGood"]["good"]))
            data.append(recipeAsset["producedGood"]["amount"])
            data[1] = translateLang(recipeAsset["producedGood"]["good"]) +" T"+str(grade)
        elif "refGood" in recipeAsset:
            data.append(translateLang(recipeAsset["refGood"]["good"]))
            data.append(recipeAsset["refGood"]["amount"])
            data[1] = translateLang(recipeAsset["refGood"]["good"]) +" T"+str(grade)
        else: 
            data.append(None)
            data.append(None)
        if "productionTime" in recipeAsset and data[-1] is not None:
            data.append(data[-1]/recipeAsset["productionTime"]*60*4)
            data.append(60*4/recipeAsset["productionTime"])
        else:
            data.append(None)
            data.append(None)
        if recipeAsset["m_Name"]=="- todelete - Pack of Luxury Goods":
            data[1] = "已被删除的配方"
        filled=0
        if "requiredGoods" in recipeAsset:
            filled = len(recipeAsset["requiredGoods"])
            required = recipeAsset["requiredGoods"]
            if isinstance(required,dict):
                requireGoodInfo = recipeAsset["requiredGoods"]
                lines = max(lines,len(requireGoodInfo["goods"]))
                data.append(goodsInfoStr(requireGoodInfo["goods"]))
            else:
                for requireGoodInfo in recipeAsset["requiredGoods"]:
                    lines = max(lines,len(requireGoodInfo["goods"]))
                    data.append(goodsInfoStr(requireGoodInfo["goods"]))
        for i in range(4-filled):
            data.append("")
        data.append("\n".join([translateLang(x) for x in recipeAsset["tags"]]))
        dataLists.append(data)
        recipeNames[recipeAssetName] = data[1]
        linesHeightInfo.append(lines)
        pc.tick()
    pc.finish()

    ov_sheet = workbook.create_sheet(("配方")[0:30])
    ov_sheet["A1"] = "配方一览"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 93
    ov_sheet["A2"] = """这里枚举了所有workshop的配方，
        采集、服务建筑、种植配方不在这个表格中
        """
    
    nameList = [
        "程序名称（Unity Prefab名称）",
        "配方名称",
        "配方类别",
        "等级",
        "生产时间",
        "产出物品",
        "产出数量",
        "每季度产出",
        "每季度生产次数",
        "需求物品",
        "",
        "",
        "",
        "标签"
    ]
    width = [
        5,
        20,
        10,
        10,
        10,
        15,
        8,
        10,
        15,
        15,
        15,
        15,
        15,
        30
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)

    frow = ov_sheet.max_row
    
    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    alignment = Alignment(wrap_text=True,vertical="center",horizontal="center")
    for i in range(4,frow):
        if linesHeightInfo[i-4]<=3:
            ov_sheet.row_dimensions[i].height=40
            
    for row in ov_sheet.iter_rows(min_row=4, max_row=frow, min_col=2, max_col=13):
        for cell in row:
            cell.alignment = alignment
    for cell in ov_sheet["H"]:
        cell.number_format = "0.00"
    for cell in ov_sheet["I"]:
        cell.number_format = "0.00"
    for cell in ov_sheet["D"]:
        cell.alignment = Alignment(vertical="center",horizontal="center")
    addEnumColor(ov_sheet,'N4:N%d'%frow)
    addEnumColor(ov_sheet,'F4:F%d'%frow)
    addEnumColor(ov_sheet,'C4:C%d'%frow)
    addEnumColor(ov_sheet,'D4:D%d'%frow,None,{
        "☆":"deffc0",
        "★":"c0ffc8",
        "★★":"a2ffae",
        "★★★":"7eff8e",
        "★★★★":"7eff8e"
    })
    ov_sheet.conditional_formatting.add('E4:E%d'%frow, data_bar_rule)
    ov_sheet.conditional_formatting.add('H4:H%d'%frow, data_bar_rule)
    ov_sheet.conditional_formatting.add('I4:I%d'%frow, data_bar_rule)
    print("Generated recipes, row to %d"%frow)
    
    #--------------------------------------------

    dataLists=[]
    linesHeightInfo=[]
    pc = ProgressCounter("Building",len(buildingsGather))
    for buildingAssetName, buildingAsset in buildingsGather.items():
        data =[
            buildingAsset["m_Name"],
            translateLang(buildingAsset["m_Name"]),
            translateLang(buildingAsset["category"])
        ]
        data.append(tryAdd(buildingAsset,"workplaces",len))
        data.append(tryAdd(buildingAsset,"maxStorage"))
        if "recipes" in buildingAsset:
            names=[]
            for recipeName in buildingAsset["recipes"]:
                if recipeName not in recipeNames:
                    names.append(translateLang(recipeName))
                else:
                    names.append(recipeNames[recipeName])
            data.append("\n".join(names))
        else:
            data.append(None)
        data.append(tryAdd(buildingAsset,"profession",translateLang))
        data.append("\n".join(translateLang(x) for x in buildingAsset["tags"]))
        data.append("\n".join(translateLang(x) for x in buildingAsset["usabilityTags"]))
        data.append(tryAdd(buildingAsset,"requiredGoods",goodsInfoStr))
        data.append(tryAdd(buildingAsset,"constructionPerSec",lambda x:1.0/x))
        data.append(tryAdd(buildingAsset,"maxBuilders"))
        data.append(buildingAsset["footprintMap"]["width"])
        data.append(buildingAsset["footprintMap"]["height"])
        data.append(bool2str(buildingAsset["movable"]))
        data.append(tryAdd(buildingAsset,"movingCost",goodInfoStr))
        data.append(bool2str(buildingAsset["destroyable"]))
        data.append(buildingAsset["refundMaterials"])
        data.append(bool2str(buildingAsset["canRotate"]))
        data.append(bool2str(buildingAsset["traversable"]))
        data.append(bool2str(buildingAsset["repeatable"]))
        data.append(tryAdd(buildingAsset,"cystsAmount"))
        data.append(tryAdd(buildingAsset,"levels",len))
        data.append(tryAdd(buildingAsset,"progressScore"))
        deco = tryAdd(buildingAsset,"hasDecorationTier")
        if deco is None:
            data.append(None)
            data.append(None)
        else:
            data.append(tryAdd(buildingAsset,"tier",translateLang))
            data.append(tryAdd(buildingAsset,"decorationScore"))
        
        data.append(bool2str(buildingAsset["canBeRuined"]))
        data.append(bool2str(buildingAsset["dlc"]))
        data.append(translateLang(buildingAsset["description"]["key"]))

            
        dataLists.append(data)
        pc.tick()
    pc.finish()

    ov_sheet = workbook.create_sheet(("建筑")[0:30])
    ov_sheet["A1"] = "所有建筑"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 93
    ov_sheet["A2"] = """这里枚举了所有建筑，包括自然生成的建筑
        """
    
    nameList = [
        "程序名称（Unity Prefab名称）",
        "建筑名称",
        "类别",
        "工作人数",
        "建筑存储",
        "生产配方",
        "专业化标签",
        "舒适标签",
        "标签",
        "建设材料",
        "建设时间",
        "建筑工上限",
        "宽x",
        "长y",
        "能移动",
        "移动成本",
        "能摧毁",
        "返还比例",
        "可旋转",
        "可穿过",
        "可重复",
        "囊肿上限",
        "最高等级",
        "城市分",
        "装饰等级",
        "装饰分",
        "可以成为废墟",
        "DLC限定",
        "描述"
    ]
    width = [
        5,
        25,
        15,
        8,
        8,
        20,
        12,
        12,
        15,
        20,
        8,
        8,
        6,
        6,
        6,
        10,
        6,
        6,
        6,
        6,
        6,
        8,
        8,
        6,
        15,
        6,
        12,
        8,
        100
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)

    frow = ov_sheet.max_row
    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    alignment = Alignment(wrap_text=True,vertical="center",horizontal="center")
    for row in ov_sheet.iter_rows(min_row=4, max_row=frow, min_col=2, max_col=30):
        for cell in row:
            cell.alignment = alignment
    
    for cell in ov_sheet["K"]:
        cell.number_format = "0.0"
    ov_sheet.conditional_formatting.add('O4:O%d'%frow, rule_tick)
    ov_sheet.conditional_formatting.add('O4:O%d'%frow, rule_cross)
    ov_sheet.conditional_formatting.add('Q4:Q%d'%frow, rule_tick)
    ov_sheet.conditional_formatting.add('Q4:Q%d'%frow, rule_cross)
    ov_sheet.conditional_formatting.add('S4:S%d'%frow, rule_tick)
    ov_sheet.conditional_formatting.add('S4:S%d'%frow, rule_cross)
    ov_sheet.conditional_formatting.add('T4:T%d'%frow, rule_tick)
    ov_sheet.conditional_formatting.add('T4:T%d'%frow, rule_cross)
    ov_sheet.conditional_formatting.add('U4:U%d'%frow, rule_tick)
    ov_sheet.conditional_formatting.add('U4:U%d'%frow, rule_cross)
    ov_sheet.conditional_formatting.add('AA4:AA%d'%frow, rule_tick)
    ov_sheet.conditional_formatting.add('AA4:AA%d'%frow, rule_cross)
    ov_sheet.conditional_formatting.add('AB4:AB%d'%frow, rule_tick)
    ov_sheet.conditional_formatting.add('AB4:AB%d'%frow, rule_cross)
    ov_sheet.conditional_formatting.add('D4:D%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('E4:E%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('L4:L%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('V4:V%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('X4:X%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('R4:R%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('K4:K%d'%frow, data_bar_rule)
    addEnumColor(ov_sheet,'C4:C%d'%frow)
    addEnumColor(ov_sheet,'G4:G%d'%frow)
    addEnumColor(ov_sheet,'H4:H%d'%frow)
    addEnumColor(ov_sheet,'I4:I%d'%frow)
    addEnumColor(ov_sheet,'G4:G%d'%frow)
    addEnumColor(ov_sheet,'Y4:Y%d'%frow)
    addEnumColor(ov_sheet,'Z4:Z%d'%frow)
    addEnumColor(ov_sheet,'X4:X%d'%frow)

    # -------------------------------------------------

def arrayToText(listValues):
    if listValues is None or len(listValues)==0:
        return None
    return "\n".join(listValues)

def arrayToTextProcessed(func,listValues):
    if listValues is None or len(listValues)==0:
        return None
    return "\n".join([func(x) for x in listValues])

def arrayToTranslationText(listValues):
    if listValues is None or len(listValues)==0:
        return None
    return "\n".join([translateLang(x) for x in listValues])

def GenBlueprintSheet(workbook:openpyxl.Workbook):
    dataLists=[]
    
    pc = ProgressCounter("Blueprint",len(effectsGather))
    for blueprintConfigName,blueprintConfig in blueprintConfigsGather.items():
        basedata=[
            blueprintConfig["m_Name"],
            translateLang(blueprintConfig["m_Name"]),
            bool2str(blueprintConfig["allowDuplicates"]),
            blueprintConfig["forcedWildcards"],
            blueprintConfig["rerollCost"]["amount"],
            blueprintConfig["rerollCostIncrease"]
        ]
        for blueprintGroup in blueprintConfig["blueprints"]:
            basedata2 = [x for x in basedata]
            basedata2+=[
                "基础池",
                blueprintGroup["range"]["x"],
                blueprintGroup["range"]["y"],
            ]
            weights = blueprintGroup["weight"]
            chance = blueprintGroup["chance"]
            for i,blueprintSetName in enumerate(blueprintGroup["sets"]):
                data = [x for x in basedata2]
                data+=[
                    weights[i],
                    chance[i],
                    translateLang(blueprintSetName)
                ]
                dataLists.append(data)
        wildCardGroup = blueprintConfig["wildcardGroup"]
        if wildCardGroup is not None:
            data = [x for x in basedata]
            data+=[
                "野池",
                None,
                None,
                None,
                None,
                translateLang(blueprintConfig["wildcardGroup"])
            ]
            dataLists.append(data)
        pc.tick()
    pc.finish()

    ov_sheet = workbook.create_sheet(("蓝图总配置")[0:30])
    ov_sheet["A1"] = "蓝图配置组 - 组别配置信息"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 120
    ov_sheet["A2"] = """这里是配置组，每场游戏开始时，会使用其中一个配置作为蓝图生成
所有常规游戏使用[默认]配置，威望难度则会使用[进阶]配置，沿海林地比较特殊，拥有单独的配置
<声望值范围> 声望值将决定从哪些[基础池]中选出本次使用的蓝图池，每一组声望配置提供一组候选蓝图池
<蓝图池名称> 决定池子后，使者会从这个池子中抽取蓝图作为最终的候选蓝图
<蓝图池类型> 分为[基础池]和[野池]，基础池最多提供2个蓝图选项，剩下的则由[野池]提供
<最低野卡组抽取数> 最终蓝图选项奖励中，至少有多少个选项来自[野池]
        """
    
    nameList = [
        "程序名称（Unity Prefab名称）",
        "蓝图总配置名称",
        "重复蓝图",
        "最低野池抽取数",
        "重选花费",
        "花费增长",

        "蓝图池类型",
        "声望值范围",
        "",
        "权重",
        "组内概率",
        "蓝图池名称"

    ]
    width = [
        5,
        30,
        8,
        9,
        9,
        9,
        11,
        6,
        6,
        6,
        10,
        30
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)

    frow = ov_sheet.max_row
    
    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    
    for cell in ov_sheet["K"]:
        cell.number_format = "0.00%"
    ov_sheet.conditional_formatting.add('C4:C%d'%frow, rule_tick)
    ov_sheet.conditional_formatting.add('C4:C%d'%frow, rule_cross)
    ov_sheet.conditional_formatting.add('J4:J%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('K4:K%d'%frow, data_bar_rule)
    ov_sheet.conditional_formatting.add('H4:I%d'%frow, color_better_dark_rule)
    ov_sheet.conditional_formatting.add('E4:F%d'%frow, color_worse_rule)
    #ov_sheet.conditional_formatting.add('D4:D%d'%frow, color_worse_rule)
    addEnumColor(ov_sheet,"B4:B%d"%frow)
    addEnumColor(ov_sheet,"G4:G%d"%frow)
    #addEnumColor(ov_sheet,"H4:I%d"%frow)

    #-------------------------------------

    dataLists=[]
    for blueprintGenName,blueprintGen in blueprintGenGather.items():
        basedata=[
            blueprintGen["m_Name"],
            translateLang(blueprintGen["m_Name"]),
            "野池" if "isWild" in blueprintGen else "基础池"
        ]
        for building in blueprintGen["buildings"]:
            data = [x for x in basedata]
            data+=[
                building["weight"],
                building["chance"],
                translateLang(building["building"])
            ]
            dataLists.append(data)


    ov_sheet = workbook.create_sheet(("蓝图池")[0:30])
    ov_sheet["A1"] = "蓝图池"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 93
    ov_sheet["A2"] = """这里描述了每个蓝图池里有什么
注意每个蓝图池内部的蓝图都有权重
        """
    
    nameList = [
        "程序名称（Unity Prefab名称）",
        "蓝图池名称",
        "蓝图池类型",
        "权重",
        "组内概率",
        "建筑蓝图名称"

    ]
    width = [
        5,
        30,
        14,
        8,
        10,
        30
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)

    frow = ov_sheet.max_row
    
    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    
    for cell in ov_sheet["E"]:
        cell.number_format = "0.00%"
    ov_sheet.conditional_formatting.add('D4:D%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('E4:E%d'%frow, data_bar_rule)
    addEnumColor(ov_sheet,"B4:B%d"%frow)
    addEnumColor(ov_sheet,"C4:C%d"%frow)
    addEnumColor(ov_sheet,"F4:F%d"%frow)
#-------------------------------------------------------------------------------------------
    

class ProgressCounter:

    def __init__(self,name,maxValue):
        self.name=name
        self.count=0
        self.maxV=maxValue
        self.startTime = time.time()

    def tick(self):
        self.count+=1
        if time.time()-self.startTime<0.5:
            return
        percent = self.count/self.maxV*100.0
        barLength=10
        fillCount = int(percent/barLength)
        nonFillCount = barLength- fillCount
        bar = "■"*fillCount+" "*nonFillCount
        print("%s [ %5d | %5d ] %7.2f%% [%s]"%(self.name,self.count,self.maxV,percent,bar),end="\r")
        self.startTime = time.time()
    
    def finish(self):
        print("\nFinish %s!"%self.name)

def GenEffectsSheet(workbook:openpyxl.Workbook):
    dataLists=[]
    pc = ProgressCounter("Effect",len(effectsGather))
    removeAssetSuffixFunc = partial(arrayToTextProcessed,lambda x: x.replace(".asset",""))

    rarityMap={
        0:"无稀有度",
        1:"白色普通",
        2:"绿色罕见",
        3:"蓝色稀有",
        4:"紫色史诗",
        5:"橙色传说",
        6:"雷铸神秘"
    }
    rarityColorMap={
        0:"aaaaaa",
        1:"efefef",
        2:"1cff00",
        3:"006ddd",
        4:"a033ed",
        5:"ff7f00",
        6:"ffe83a"
    }
    rarityNameColorMap={}
    for k,v in rarityColorMap.items():
        rarityNameColorMap[rarityMap[k]]=v
    def rarityMapping(rarity:int):
        return rarityMap[rarity]
    for effectFileNameKey,effect in effectsGather.items():
        data=[
            effect["m_Name"],
            translateLang(effect["m_Name"]),
        ]
        if "displayName" in effect:
            data.append(translateLang(effect["displayName"]["key"]))
        else:
            data.append(None)
        data.append(tryAdd(effect,"rarity",rarityMapping))
        data.append(tryAdd(effect,"isPerk",bool2str,0))
        data.append(tryAdd(effect,"isEthereal",bool2str,0))
        if "description" in effect:
            data.append(translateLang(effect["description"]["key"]))
        else:
            data.append(None)
        reward = tryAdd(effect,"rewards",partial(arrayToTextProcessed,lambda x: translateLang(x.replace(".asset","")).replace(".asset","")))
        if reward is None: # HookedEffectModel
            reward = tryAdd(effect,"hookedEffects",partial(arrayToTextProcessed,lambda x: "[条件]"+translateLang(x.replace(".asset","")).replace(".asset","")))
            if reward is not None:
                reward+="\n"
                reward += filterNoneToStr(tryAdd(effect,"instantEffects",partial(arrayToTextProcessed,lambda x: "[立即]:"+translateLang(x.replace(".asset","")).replace(".asset",""))))
        data.append(reward)
        data.append(tryAdd(effect,"tradingBuyValue"))
        data.append(tryAdd(effect,"isPositive",bool2str))
        data.append(tryAdd(effect,"drawLimit", lambda x: None if x==0 else x))
        data.append(tryAdd(effect,"blockedBy",partial(arrayToTextProcessed,lambda x: translateLang(x.replace(".asset","")).replace(".asset",""))))
        data.append(tryAdd(effect,"dlc",bool2str))
        data.append(tryAdd(effect,"label",lambda x:x.replace(".asset","")))
        data.append(tryAdd(effect,"usabilityTags",arrayToTranslationText))
        data.append(tryAdd(effect,"m_Script"))
        dataLists.append(data)
        pc.tick()
    pc.finish()


    ov_sheet = workbook.create_sheet(("效果")[0:30])
    ov_sheet["A1"] = "效果 Effects"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 120
    ov_sheet["A2"] = """游戏里的所有基石，辅助技能，升级之类的，几乎都是通过Effects实现的
这个表格可能对modder非常有帮助_(:з」∠)_
表格会包含大量基础的程序信息，推荐给想写基石的人用
注意大部分效果是<混合效果> CompositeEffectModel [条件]意思是触发条件后才添加效果，[立即]则是获得时立刻获得相应效果
另外标准价值不是琥珀价值，琥珀出售价值初始为12，购买价值初始为15
这里所有效果名称都是用程序名+自制简单程序名机翻，文本翻译会有重复和未翻译的情况
        """
    
    nameList = [
        "程序名称（Unity Prefab名称）",
        "程序名称翻译",
        "中文名称",
        "稀有度",
        "是基石",
        "仅限暂停",
        "效果描述",
        "实际效果列表",
        "标准价值",
        "正面效果",
        "数量限制",
        "致其不出现的基石",
        "DLC限定",
        "类型标签",
        "可用性标签",
        "效果代码文件",
    ]
    width = [
        25,
        20,
        15,
        10,
        8,
        8,
        35,
        25,
        10,
        8,
        8,
        25,
        8,
        25,
        15,
        20
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)

    frow = ov_sheet.max_row
    
    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    
    alignment = Alignment(wrap_text=True,vertical="center",horizontal="center")
    alignmentCenterOnly = Alignment(vertical="center",horizontal="center")
    
    for row in ov_sheet.iter_rows(min_row=4, max_row=frow, min_col=1, max_col=16):
        for cell in row:
            cell.alignment = alignmentCenterOnly
    for row in ov_sheet.iter_rows(min_row=4, max_row=frow, min_col=7, max_col=8):
        for cell in row:
            cell.alignment = alignment
    for row in ov_sheet.iter_rows(min_row=4, max_row=frow, min_col=15, max_col=15):
        for cell in row:
            cell.alignment = alignment
    #for cell in ov_sheet["K"]:
    #    cell.number_format = "0.00%"
    addTickCrossColor(ov_sheet,'E4:E%d'%frow)
    addTickCrossColor(ov_sheet,'F4:F%d'%frow)
    addTickCrossColor(ov_sheet,'J4:J%d'%frow)
    addTickCrossColor(ov_sheet,'M4:M%d'%frow)
    #ov_sheet.conditional_formatting.add('J4:J%d'%frow, color_compare_rule)
    #ov_sheet.conditional_formatting.add('K4:K%d'%frow, data_bar_rule)
    #ov_sheet.conditional_formatting.add('H4:I%d'%frow, color_better_dark_rule)
    #ov_sheet.conditional_formatting.add('E4:F%d'%frow, color_worse_rule)
    ov_sheet.conditional_formatting.add('I4:I%d'%frow, data_bar_rule)
    ov_sheet.conditional_formatting.add('K4:K%d'%frow, data_bar_rule)
    #ov_sheet.conditional_formatting.add('C4:C%d'%frow, color_better_dark_rule)
    addEnumColor(ov_sheet,"D4:D%d"%frow,[],rarityNameColorMap)
    addEnumColor(ov_sheet,"O4:O%d"%frow)
    addEnumColor(ov_sheet,"N4:N%d"%frow)
    addEnumColor(ov_sheet,"P4:P%d"%frow)


    #--------------------------------------------------

    dataLists = []
    
    pc = ProgressCounter("EffectsTable General",len(effectsTablesGather))
    totalEffects = 0
    for effectsTableKey, effects in effectsTablesGather.items():
        data = [
            effects["m_Name"],
            translateLang(effects["m_Name"]),
            effects["amounts"]["x"],
            effects["amounts"]["y"],
            len(effects["guaranteedEffects"]),
            len(effects["effects"])
        ]
        totalEffects+=len(effects["guaranteedEffects"]) + len(effects["effects"])
        dataLists.append(data)
        pc.tick()
        pass
    pc.finish()

    ov_sheet = workbook.create_sheet(("战利品表-总览")[0:30])
    ov_sheet["A1"] = "战利品表 Effect Tables"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 120
    ov_sheet["A2"] = """游戏里的<森林事件奖励><每年发的基石> 都是根据战利品表决定的
每次决定<事件抉择奖励>和<女王n选1基石>时，都会从相应的战利品表抽取相应数量的效果
具体列表参考<战利品表-效果>，这个列表只提供总览信息
        """
    
    nameList = [
        "程序名称（Unity Prefab名称）",
        "战利品表名翻译",
        "最小数量",
        "最大数量",
        "必定出现效果数",
        "随机池效果数"
    ]
    width = [
        35,
        30,
        12,
        12,
        12,
        12
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)

    frow = ov_sheet.max_row
    
    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    
    alignment = Alignment(wrap_text=True,vertical="center",horizontal="center")
    alignmentCenterOnly = Alignment(vertical="center",horizontal="center")
    
    #for row in ov_sheet.iter_rows(min_row=4, max_row=frow, min_col=1, max_col=16):
    #    for cell in row:
    #        cell.alignment = alignmentCenterOnly
    #for row in ov_sheet.iter_rows(min_row=4, max_row=frow, min_col=7, max_col=8):
    #    for cell in row:
    #        cell.alignment = alignment
    #for row in ov_sheet.iter_rows(min_row=4, max_row=frow, min_col=15, max_col=15):
    #    for cell in row:
    #        cell.alignment = alignment
    #for cell in ov_sheet["K"]:
    #    cell.number_format = "0.00%"
    addTickCrossColor(ov_sheet,'E4:E%d'%frow)
    addTickCrossColor(ov_sheet,'F4:F%d'%frow)
    addTickCrossColor(ov_sheet,'J4:J%d'%frow)
    addTickCrossColor(ov_sheet,'M4:M%d'%frow)
    #ov_sheet.conditional_formatting.add('J4:J%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('C4:C%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('D4:D%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('E4:E%d'%frow, color_better_rule)
    ov_sheet.conditional_formatting.add('F4:F%d'%frow, data_bar_rule)
    #addEnumColor(ov_sheet,"D4:D%d"%frow,[],rarityNameColorMap)

    #-------------------------------------

    dataLists=[]
    pc = ProgressCounter("EffectsTable Detail",totalEffects)
    for effectsTableKey, effects in effectsTablesGather.items():
        dataBase = [
            translateLang(effects["m_Name"]),
        ]

        for guaranteedEffect in effects["guaranteedEffects"]:
            data = [x for x in dataBase]
            data += ["√",
                    None,
                    1.0,
                    translateLang(guaranteedEffect.replace(".asset","")),
                    guaranteedEffect
                    ]
            effectObj = effectsGather[guaranteedEffect]
            if "description" in effectObj:
                data.append(translateLang(effectObj["description"]["key"]))
            else:
                data.append(None)
            dataLists.append(data)
            pc.tick()
        for effect in effects["effects"]:
            data = [x for x in dataBase]
            data += ["×",
                    effect["chance"],
                    effect["actualChance"],
                    translateLang(effect["effect"].replace(".asset","")),
                    effect["effect"]
                    ]
            effectObj = effectsGather[effect["effect"]]
            if "description" in effectObj:
                data.append(translateLang(effectObj["description"]["key"]))
            else:
                data.append(None)
            dataLists.append(data)
            pc.tick()
        pass
    pc.finish()

    ov_sheet = workbook.create_sheet(("战利品表-细节")[0:30])
    ov_sheet["A1"] = "战利品表 Effect Tables"
    ov_sheet.merge_cells('A2:P2')
    ov_sheet["A2"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ov_sheet.row_dimensions[2].height = 120
    ov_sheet["A2"] = """游戏里的<森林事件奖励><每年发的基石> 都是根据战利品表决定的
每次决定<事件抉择奖励>和<女王n选1基石>时，都会从相应的战利品表抽取相应数量的效果，并加权抽取
这个表格展示了所有战利品表的权重
        """
    
    nameList = [
        "战利品表名翻译",
        "必定获得",
        "权重",
        "组内概率",
        "效果名称",
        "效果程序名",
        "效果描述"
    ]
    width = [
        30,
        8,
        12,
        12,
        35,
        35,
        70
    ]
    _addTitleBar(nameList,width,ov_sheet,1,3)
    _addDataToWorkSheet(dataLists,ov_sheet,1,4)

    frow = ov_sheet.max_row
    
    for cell in ov_sheet[1]:
        cell.font = bold_font
    for cell in ov_sheet[3]:
        cell.font = bold_font
    
    alignment = Alignment(wrap_text=True,vertical="center",horizontal="center")
    alignmentCenterOnly = Alignment(vertical="center",horizontal="center")
    
    #for row in ov_sheet.iter_rows(min_row=4, max_row=frow, min_col=1, max_col=16):
    #    for cell in row:
    #        cell.alignment = alignmentCenterOnly
    #for row in ov_sheet.iter_rows(min_row=4, max_row=frow, min_col=7, max_col=8):
    #    for cell in row:
    #        cell.alignment = alignment
    #for row in ov_sheet.iter_rows(min_row=4, max_row=frow, min_col=15, max_col=15):
    #    for cell in row:
    #        cell.alignment = alignment
    for cell in ov_sheet["D"]:
        cell.number_format = "0.00%"
    addTickCrossColor(ov_sheet,'B4:B%d'%frow)
    ov_sheet.conditional_formatting.add('C4:C%d'%frow, color_compare_rule)
    ov_sheet.conditional_formatting.add('D4:D%d'%frow, data_bar_rule)
    addEnumColor(ov_sheet,"E4:E%d"%frow)
    addEnumColor(ov_sheet,"F4:F%d"%frow)
    addEnumColor(ov_sheet,"A4:A%d"%frow)



#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------

buildingsGenGather = loadJson("output/buildings_gen.json")
depositsGather = loadJson("output/deposites_gen.json")
gladesGather = loadJson("output/glades.json")
gladesGenGather = loadJson("output/glades_gen_model.json")
relicGenGather = loadJson("output/relics_gen.json")
biomesGather = loadJson("output/biomes.json")
oresGenGather = loadJson("output/ores_gen.json")
springsGenGather = loadJson("output/springs_gen.json")
spGladeGenGather = loadJson("output/glades_sp_group.json")
displayNamesGather = loadJson("output/display_names.json")
difficultiesGather = loadJson("output/difficulties.json")
relicsGather = loadJson("output/relics.json")
effectsGather = loadJson("output/effects.json")
effectsTablesGather = loadJson("output/effects_table.json")
goodsGather = loadJson("output/goods.json")
recipesGather = loadJson("output/recipes.json")
buildingsGather = loadJson("output/buildings.json")
blueprintConfigsGather = loadJson("output/blueprints_configs.json")
blueprintGenGather = loadJson("output/blueprints_gen.json")
effectsGather = loadJson("output/effects.json")
globalSortedDisplayName = [(k.replace(".asset",""),v) for k,v in displayNamesGather.items()]
globalSortedDisplayName = sorted(globalSortedDisplayName,key=lambda x:len(x[0]),reverse=True)


for k,v in effectsTablesGather.items():
    effects=v["effects"]
    totalWeight=0
    for effectInfo in effects:
        totalWeight+=effectInfo["chance"]
    for effectInfo in effects:
        effectInfo["actualChance"] = effectInfo["chance"]/totalWeight

for k,v in blueprintConfigsGather.items():
    blueprintsGroupGroup = v["blueprints"]
    for blueprintGroup in blueprintsGroupGroup:
        weights= []
        totalWeight=0
        for blueprintSetName in blueprintGroup["sets"]:
            w=blueprintGenGather[blueprintSetName]["weight"]
            weights.append(w)
            totalWeight+=w
        chance=[]
        for i in range(len(weights)):
            chance.append(weights[i]/totalWeight)
        blueprintGroup["weight"]=weights
        blueprintGroup["chance"]=chance
    wildcardGroup = v["wildcards"]
    if len(wildcardGroup)>0:
        name = k.replace(".asset","") + "WildCard"
        obj={}
        obj["m_Script"]="BuildingsWeightedContainer.cs"
        obj["m_Name"]=name
        obj["weight"]=None
        obj["buildings"]=v["wildcards"]
        obj["isWild"]=True
        blueprintGenGather[name] = obj
        v["wildcardGroup"]=name
    else:
        v["wildcardGroup"] = None

for k,v in blueprintGenGather.items():
    totalWeight=0
    for building in v["buildings"]:
        totalWeight += building["weight"]
    for building in v["buildings"]:
        building["chance"] = building["weight"] / totalWeight

#pre process
for k,v in gladesGather.items():
    v["as_init_glade"] = 0
    v["as_main_glade"] = 0
    v["as_sub_glade"] = 0
    v["as_sp_glade"] = 0
    v["appear_in_gen_model"] = set()
    v["total_weight"] = 0
    if "springs" not in v:
        v["springs"]=[]
    if "ore" not in v:
        v["ore"] = []
for k,v in gladesGenGather.items():
    initGlades = None
    if "initialGlades" in v: # older version
        initGlades=v["initialGlades"]
        for glade in initGlades:
            gladesGather[glade]["as_init_glade"]+=1
            gladesGather[glade]["appear_in_gen_model"].add(k)
    else:
        initGlades=v["initial"] #v1.4 bay generation code
        v["initialGlades"] = []
        for gladeInfo in initGlades:
            glade = gladeInfo["glade"]
            gladesGather[glade]["as_init_glade"]+=1
            gladesGather[glade]["appear_in_gen_model"].add(k)
            v["initialGlades"].append(glade)
    for level in v["levels"]:
        for glade in level["glades"]:
            gladeName = glade["glade"]
            gladesGather[gladeName]["as_main_glade"]+=1
            gladesGather[gladeName]["appear_in_gen_model"].add(k)
            gladesGather[gladeName]["total_weight"]+=glade["weight"]
    for levelBase in v["extraIterations"]:
        level = levelBase["iteration"]
        for glade in level["glades"]:
            gladeName = glade["glade"]
            gladesGather[gladeName]["as_sub_glade"]+=1
            gladesGather[gladeName]["appear_in_gen_model"].add(k)
            gladesGather[gladeName]["total_weight"]+=glade["weight"]

for k,v in spGladeGenGather.items():
    for gladeInfo in v["glades"]:
        glade = gladeInfo["glade"]
        gladesGather[glade]["appear_in_gen_model"].add(k)
        gladesGather[glade]["as_sp_glade"] +=1

for name,relicDicts in relicGenGather.items():
    for level,relicList in relicDicts.items():
        totalWeight = 0
        for relicInfo in relicList:
            if relicInfo["forceUniqueness"]==0:
                totalWeight+=relicInfo["weight"]
        #countRelics = Counter([relicInfo["relic"] for relicInfo in relicList])
        for relicInfo in relicList:
            if relicInfo["forceUniqueness"]==0:
                relicInfo["groupChance"]=relicInfo["weight"]/totalWeight
                #relicInfo["max"] = countRelics[relicInfo["relic"]]
                #relicInfo["min"] = 0
            else:
                relicInfo["groupChance"]=relicInfo["chance"]
                #relicInfo["max"] = 1
                #relicInfo["min"] = 0


for name,depositDicts in depositsGather.items():
    for level,depositList in depositDicts.items():
        totalWeight = 0
        for depositInfo in depositList:
            totalWeight+=depositInfo["amount"]
        for depositInfo in depositList:
            depositInfo["groupChance"]=depositInfo["amount"]/totalWeight

for name,buildingDicts in buildingsGenGather.items():
    for level,buildingList in buildingDicts.items():
        totalWeight = 0
        for buildingInfo in buildingList:
            totalWeight+=buildingInfo["weight"]
        for buildingInfo in buildingList:
            buildingInfo["groupChance"]=buildingInfo["weight"]/totalWeight

for name,oreDicts in oresGenGather.items():
    for oreLists in oreDicts:
        oreList=oreLists["chances"]
        totalWeight = 0
        for oreInfo in oreList:
            totalWeight+=oreInfo["amount"]
        for oreInfo in oreList:
            oreInfo["groupChance"]=oreInfo["amount"]/totalWeight

springsGatherDic = {}
for typeName,listInfo in springsGenGather.items():
    dic={}
    springsGatherDic[typeName] = dic
    for info in listInfo:
        lv = info["level"]
        if lv not in dic:
            dic[lv]=[]
        dic[lv].append(info)
    for index, springList in dic.items():
        totalWeight = 0
        for spring in springList:
            totalWeight+=spring["weight"]
        for spring in springList:
            spring["groupChance"] = spring["weight"]/totalWeight


# main part

def buildExcelGlades():
    wb = openpyxl.load_workbook("template_glades.xlsx")
    processGenGladeModelIntegrate(wb,gladesGenGather,spGladeGenGather)
    processGlade(wb,gladesGather)
    wb.save("vis_output/空地模板.xlsx")

def buildExcelGoods():
    wb = openpyxl.load_workbook("template_good.xlsx")
    GenGoodSheet(wb)
    wb.save("vis_output/物品配方与建筑.xlsx")

def buildExcelBlueprint():
    wb = openpyxl.load_workbook("template_blueprint.xlsx")
    GenBlueprintSheet(wb)
    wb.save("vis_output/蓝图机制详解.xlsx")

def buildEffects():
    wb = openpyxl.load_workbook("template_effects.xlsx")
    GenEffectsSheet(wb)
    wb.save("vis_output/基石与效果.xlsx")
    pass
buildExcelGlades()
buildExcelGoods()
buildExcelBlueprint()
buildEffects()